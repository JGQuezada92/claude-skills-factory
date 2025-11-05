#!/usr/bin/env python3
"""
Output Validator

Validates calculations, logical consistency, and numerical accuracy for 
liquidity analysis outputs. Provides comprehensive error checking before 
finalizing Excel models and reports.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime
import re
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ValidationResult:
    """Container for validation results"""
    
    def __init__(self):
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, Any]] = []
        self.passed: List[Dict[str, Any]] = []
        self.is_valid = True
    
    def add_error(self, check_name: str, message: str, details: Dict = None):
        """Add an error to the validation results"""
        self.errors.append({
            'check': check_name,
            'message': message,
            'details': details or {}
        })
        self.is_valid = False
    
    def add_warning(self, check_name: str, message: str, details: Dict = None):
        """Add a warning to the validation results"""
        self.warnings.append({
            'check': check_name,
            'message': message,
            'details': details or {}
        })
    
    def add_passed(self, check_name: str, message: str = "Check passed"):
        """Add a passed check"""
        self.passed.append({
            'check': check_name,
            'message': message
        })
    
    def get_summary(self) -> Dict:
        """Get summary of validation results"""
        return {
            'is_valid': self.is_valid,
            'total_errors': len(self.errors),
            'total_warnings': len(self.warnings),
            'total_passed': len(self.passed),
            'errors': self.errors,
            'warnings': self.warnings,
            'passed': self.passed
        }


class OutputValidator:
    """Comprehensive validator for liquidity analysis outputs"""
    
    def __init__(self, tolerance: float = 0.01):
        """
        Initialize validator
        
        Args:
            tolerance: Tolerance for numerical comparisons (default 0.01 = 1%)
        """
        self.tolerance = tolerance
        self.expected_ranges = {
            'growth_rate': (-50.0, 100.0),
            'cycle_length_months': (50.0, 75.0),
            'velocity': (0.5, 15.0),
            'correlation': (-1.0, 1.0),
            'percent_change': (-100.0, 500.0)
        }
    
    def validate_percent_change(
        self,
        value_current: float,
        value_previous: float,
        reported_change: float,
        tolerance: Optional[float] = None
    ) -> Tuple[bool, str]:
        """
        Verify percent change calculations
        
        Args:
            value_current: Current value
            value_previous: Previous value
            reported_change: Reported percent change
            tolerance: Optional tolerance (uses default if None)
            
        Returns:
            Tuple of (is_valid, message)
        """
        if value_previous == 0 or pd.isna(value_previous):
            return False, "Cannot calculate percent change: previous value is zero or NaN"
        
        if pd.isna(value_current) or pd.isna(reported_change):
            return False, "Cannot validate: current value or reported change is NaN"
        
        tol = tolerance or self.tolerance
        expected_change = ((value_current - value_previous) / value_previous) * 100
        difference = abs(expected_change - reported_change)
        
        if difference <= tol:
            return True, f"Percent change validated: {reported_change:.2f}% (expected: {expected_change:.2f}%)"
        else:
            return False, (
                f"Percent change mismatch: reported {reported_change:.2f}%, "
                f"expected {expected_change:.2f}% (difference: {difference:.2f}%)"
            )
    
    def validate_growth_rate_consistency(
        self,
        analysis_results: Dict,
        data: pd.DataFrame,
        value_column: str,
        date_column: str = 'date'
    ) -> ValidationResult:
        """
        Check growth rates are consistent across time periods
        
        Args:
            analysis_results: Analysis results dictionary
            data: Source data DataFrame
            value_column: Name of value column
            date_column: Name of date column
            
        Returns:
            ValidationResult object
        """
        result = ValidationResult()
        
        try:
            df = data.copy()
            df = df.sort_values(date_column).reset_index(drop=True)
            
            # Check if we have enough data
            if len(df) < 13:
                result.add_warning(
                    'growth_rate_consistency',
                    'Insufficient data for growth rate consistency check (need at least 13 periods)'
                )
                return result
            
            # Calculate YoY manually
            current_value = df[value_column].iloc[-1]
            value_12m_ago = df[value_column].iloc[-13]
            
            if value_12m_ago != 0:
                manual_yoy = ((current_value - value_12m_ago) / value_12m_ago) * 100
                
                # Calculate cumulative MoM over 12 months
                cumulative_mom = 1.0
                for i in range(-12, 0):
                    if i > -len(df):
                        prev_val = df[value_column].iloc[i-1]
                        curr_val = df[value_column].iloc[i]
                        if prev_val != 0:
                            mom_rate = 1 + (curr_val / prev_val - 1)
                            cumulative_mom *= mom_rate
                
                cumulative_yoy = (cumulative_mom - 1) * 100
                
                # Compare manual YoY with cumulative MoM
                difference = abs(manual_yoy - cumulative_yoy)
                
                if difference > 5.0:  # 5% tolerance for compounding differences
                    result.add_warning(
                        'growth_rate_consistency',
                        f'YoY growth ({manual_yoy:.2f}%) differs from cumulative MoM ({cumulative_yoy:.2f}%)',
                        {'difference': difference, 'tolerance': 5.0}
                    )
                else:
                    result.add_passed(
                        'growth_rate_consistency',
                        f'YoY and cumulative MoM are consistent (difference: {difference:.2f}%)'
                    )
            
        except Exception as e:
            result.add_error(
                'growth_rate_consistency',
                f'Error checking growth rate consistency: {str(e)}'
            )
        
        return result
    
    def validate_cycle_phase_consistency(
        self,
        cycle_analysis: Dict,
        liquidity_data: pd.DataFrame,
        value_column: str = 'liquidity_index',
        date_column: str = 'date'
    ) -> ValidationResult:
        """
        Ensure phase determination matches data trends
        
        Args:
            cycle_analysis: Cycle analysis results
            liquidity_data: Liquidity data DataFrame
            value_column: Name of value column
            date_column: Name of date column
            
        Returns:
            ValidationResult object
        """
        result = ValidationResult()
        
        try:
            if 'current_phase' not in cycle_analysis:
                result.add_warning('cycle_phase_consistency', 'No current phase information available')
                return result
            
            phase_info = cycle_analysis['current_phase']
            phase = phase_info.get('phase', 'unknown')
            
            if phase == 'unknown':
                result.add_warning('cycle_phase_consistency', 'Phase is unknown, cannot validate')
                return result
            
            df = liquidity_data.copy()
            df = df.sort_values(date_column).reset_index(drop=True)
            
            if len(df) < 6:
                result.add_warning('cycle_phase_consistency', 'Insufficient data for trend validation')
                return result
            
            # Check recent trend
            recent_trend = df[value_column].iloc[-6:].pct_change().mean()
            recent_trend_pct = recent_trend * 100
            
            # Validate phase matches trend
            if phase == 'expansion' and recent_trend_pct < -2:
                result.add_error(
                    'cycle_phase_consistency',
                    f'Expansion phase but negative trend ({recent_trend_pct:.2f}%)',
                    {'phase': phase, 'trend': recent_trend_pct}
                )
            elif phase == 'contraction' and recent_trend_pct > 2:
                result.add_warning(
                    'cycle_phase_consistency',
                    f'Contraction phase but positive trend ({recent_trend_pct:.2f}%)',
                    {'phase': phase, 'trend': recent_trend_pct}
                )
            else:
                result.add_passed(
                    'cycle_phase_consistency',
                    f'Phase "{phase}" matches trend ({recent_trend_pct:.2f}%)'
                )
            
            # Validate cycle completion percentage
            cycle_completion = phase_info.get('cycle_completion_percent', 0)
            months_elapsed = phase_info.get('months_elapsed', 0)
            cycle_length = phase_info.get('months_to_turning_point', 0) + months_elapsed if phase_info.get('months_to_turning_point') else None
            
            if cycle_length and cycle_length > 0:
                expected_completion = (months_elapsed / cycle_length) * 100
                completion_diff = abs(cycle_completion - expected_completion)
                
                if completion_diff > 5.0:
                    result.add_error(
                        'cycle_completion_consistency',
                        f'Cycle completion mismatch: {cycle_completion:.2f}% vs expected {expected_completion:.2f}%',
                        {'reported': cycle_completion, 'expected': expected_completion, 'difference': completion_diff}
                    )
                else:
                    result.add_passed('cycle_completion_consistency', 'Cycle completion percentage is consistent')
            
        except Exception as e:
            result.add_error(
                'cycle_phase_consistency',
                f'Error validating cycle phase: {str(e)}'
            )
        
        return result
    
    def validate_excel_formulas(
        self,
        excel_path: str,
        expected_formulas: Optional[Dict[str, str]] = None
    ) -> ValidationResult:
        """
        Read and verify Excel formulas
        
        Args:
            excel_path: Path to Excel file
            expected_formulas: Optional dict mapping cell references to expected formulas
            
        Returns:
            ValidationResult object
        """
        result = ValidationResult()
        
        if not OPENPYXL_AVAILABLE:
            result.add_warning('excel_formula_validation', 'openpyxl not available, skipping Excel validation')
            return result
        
        try:
            wb = load_workbook(excel_path, data_only=False)
            
            # Check for division by zero errors
            division_by_zero_found = False
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            formula = cell.value
                            
                            # Check for division by zero
                            if formula and '/0' in str(formula) or 'DIV/0' in str(formula):
                                division_by_zero_found = True
                                result.add_error(
                                    'excel_formula_validation',
                                    f'Potential division by zero in {sheet_name}!{cell.coordinate}: {formula}'
                                )
                            
                            # Check against expected formulas if provided
                            if expected_formulas and cell.coordinate in expected_formulas:
                                expected = expected_formulas[cell.coordinate]
                                if str(formula) != expected:
                                    result.add_warning(
                                        'excel_formula_validation',
                                        f'Formula mismatch in {sheet_name}!{cell.coordinate}: got {formula}, expected {expected}'
                                    )
            
            if not division_by_zero_found:
                result.add_passed('excel_formula_validation', 'No division by zero errors found')
            
        except Exception as e:
            result.add_error(
                'excel_formula_validation',
                f'Error reading Excel file: {str(e)}'
            )
        
        return result
    
    def validate_report_numerical_accuracy(
        self,
        report_text: str,
        source_data: Dict[str, Any]
    ) -> ValidationResult:
        """
        Extract numbers from report and verify against source
        
        Args:
            report_text: Report text content
            source_data: Dictionary of source calculations
            
        Returns:
            ValidationResult object
        """
        result = ValidationResult()
        
        try:
            # Extract percentage values from report
            percent_pattern = r'(\d+\.?\d*)\s*%'
            percentages = re.findall(percent_pattern, report_text)
            
            # Extract growth rate mentions
            growth_pattern = r'(?:growth|change|increase|decrease).*?(\d+\.?\d*)\s*%'
            growth_rates = re.findall(growth_pattern, report_text, re.IGNORECASE)
            
            # Validate key metrics mentioned in report
            validated_count = 0
            mismatch_count = 0
            
            for key, expected_value in source_data.items():
                if isinstance(expected_value, (int, float)):
                    # Look for this value in the report
                    value_str = f"{expected_value:.2f}"
                    if value_str in report_text or str(int(expected_value)) in report_text:
                        validated_count += 1
                    else:
                        # Check if similar value exists (within tolerance)
                        found = False
                        for extracted in percentages + growth_rates:
                            try:
                                extracted_val = float(extracted)
                                if abs(extracted_val - expected_value) < (abs(expected_value) * self.tolerance):
                                    found = True
                                    break
                            except ValueError:
                                continue
                        
                        if not found:
                            mismatch_count += 1
                            result.add_warning(
                                'report_numerical_accuracy',
                                f'Key metric "{key}" (value: {expected_value}) not found in report with expected accuracy'
                            )
            
            if mismatch_count == 0:
                result.add_passed(
                    'report_numerical_accuracy',
                    f'All {validated_count} key metrics found in report'
                )
            
        except Exception as e:
            result.add_error(
                'report_numerical_accuracy',
                f'Error validating report numerical accuracy: {str(e)}'
            )
        
        return result
    
    def validate_range(
        self,
        value: float,
        metric_type: str,
        context: str = ""
    ) -> Tuple[bool, str]:
        """
        Check value is within expected range
        
        Args:
            value: Value to check
            metric_type: Type of metric (growth_rate, cycle_length_months, velocity, correlation, percent_change)
            context: Additional context for the check
            
        Returns:
            Tuple of (is_valid, message)
        """
        if metric_type not in self.expected_ranges:
            return True, f"Unknown metric type: {metric_type}"
        
        min_val, max_val = self.expected_ranges[metric_type]
        
        if pd.isna(value):
            return False, f"Value is NaN for {metric_type}"
        
        if value < min_val or value > max_val:
            return False, (
                f"Value {value:.2f} for {metric_type} is outside expected range "
                f"[{min_val:.2f}, {max_val:.2f}]{' - ' + context if context else ''}"
            )
        
        return True, f"Value {value:.2f} for {metric_type} is within expected range"
    
    def validate_cycle_length(
        self,
        cycle_length_months: float
    ) -> Tuple[bool, str]:
        """
        Validate cycle length is reasonable (50-75 months)
        
        Args:
            cycle_length_months: Cycle length in months
            
        Returns:
            Tuple of (is_valid, message)
        """
        return self.validate_range(cycle_length_months, 'cycle_length_months', 'Liquidity cycles typically 60-65 months')
    
    def validate_correlation(
        self,
        correlation: float
    ) -> Tuple[bool, str]:
        """
        Validate correlation coefficient is between -1 and 1
        
        Args:
            correlation: Correlation coefficient
            
        Returns:
            Tuple of (is_valid, message)
        """
        return self.validate_range(correlation, 'correlation', 'Correlation must be between -1 and 1')
    
    def validate_monetary_aggregates_hierarchy(
        self,
        aggregates: Dict[str, float]
    ) -> ValidationResult:
        """
        Verify M0 ≤ M1 ≤ M2 ≤ M3 hierarchy
        
        Args:
            aggregates: Dictionary with M0, M1, M2, M3 values
            
        Returns:
            ValidationResult object
        """
        result = ValidationResult()
        
        hierarchy = ['M0', 'M1', 'M2', 'M3']
        values = {}
        
        for agg in hierarchy:
            if agg in aggregates:
                val = aggregates[agg]
                if pd.isna(val) or val <= 0:
                    result.add_error(
                        'monetary_aggregates_hierarchy',
                        f'{agg} is missing, zero, or negative: {val}'
                    )
                    return result
                values[agg] = val
        
        # Check hierarchy
        for i in range(len(hierarchy) - 1):
            current = hierarchy[i]
            next_agg = hierarchy[i + 1]
            
            if current in values and next_agg in values:
                if values[current] > values[next_agg]:
                    result.add_error(
                        'monetary_aggregates_hierarchy',
                        f'Hierarchy violation: {current} ({values[current]}) > {next_agg} ({values[next_agg]})',
                        {'current': values[current], 'next': values[next_agg]}
                    )
                else:
                    result.add_passed(
                        'monetary_aggregates_hierarchy',
                        f'{current} ({values[current]:.2f}) ≤ {next_agg} ({values[next_agg]:.2f})'
                    )
        
        if result.is_valid:
            result.add_passed('monetary_aggregates_hierarchy', 'Monetary aggregates hierarchy is correct')
        
        return result
    
    def validate_policy_stance_consistency(
        self,
        policy_stance: str,
        balance_sheet_change: float,
        yoy_change: float
    ) -> ValidationResult:
        """
        Check policy stance matches balance sheet direction
        
        Args:
            policy_stance: Policy stance (expansive, contractive, neutral, mixed)
            balance_sheet_change: Recent balance sheet change
            yoy_change: Year-over-year change percentage
            
        Returns:
            ValidationResult object
        """
        result = ValidationResult()
        
        if policy_stance == 'expansive':
            if balance_sheet_change < 0 and yoy_change < -5:
                result.add_error(
                    'policy_stance_consistency',
                    f'Policy stance is "expansive" but balance sheet is contracting (change: {balance_sheet_change:.2f}, YoY: {yoy_change:.2f}%)',
                    {'stance': policy_stance, 'change': balance_sheet_change, 'yoy': yoy_change}
                )
            else:
                result.add_passed('policy_stance_consistency', 'Expansive policy matches balance sheet growth')
        
        elif policy_stance == 'contractive':
            if balance_sheet_change > 0 and yoy_change > 5:
                result.add_warning(
                    'policy_stance_consistency',
                    f'Policy stance is "contractive" but balance sheet is expanding (change: {balance_sheet_change:.2f}, YoY: {yoy_change:.2f}%)',
                    {'stance': policy_stance, 'change': balance_sheet_change, 'yoy': yoy_change}
                )
            else:
                result.add_passed('policy_stance_consistency', 'Contractive policy matches balance sheet contraction')
        
        else:
            result.add_passed('policy_stance_consistency', f'Policy stance "{policy_stance}" is neutral or mixed')
        
        return result
    
    def comprehensive_validation(
        self,
        analysis_results: Dict,
        source_data: Optional[pd.DataFrame] = None,
        excel_path: Optional[str] = None,
        report_text: Optional[str] = None
    ) -> ValidationResult:
        """
        Run comprehensive validation on all outputs
        
        Args:
            analysis_results: Dictionary of analysis results
            source_data: Optional source data DataFrame
            excel_path: Optional path to Excel file
            report_text: Optional report text
            
        Returns:
            ValidationResult object with all validation checks
        """
        result = ValidationResult()
        
        # Validate cycle phase consistency if available
        if 'current_phase' in analysis_results or 'cycles' in analysis_results:
            cycle_result = self.validate_cycle_phase_consistency(
                analysis_results,
                source_data if source_data is not None else pd.DataFrame()
            )
            result.errors.extend(cycle_result.errors)
            result.warnings.extend(cycle_result.warnings)
            result.passed.extend(cycle_result.passed)
        
        # Validate Excel formulas if provided
        if excel_path:
            excel_result = self.validate_excel_formulas(excel_path)
            result.errors.extend(excel_result.errors)
            result.warnings.extend(excel_result.warnings)
            result.passed.extend(excel_result.passed)
        
        # Validate report numerical accuracy if provided
        if report_text and source_data is not None:
            report_result = self.validate_report_numerical_accuracy(report_text, analysis_results)
            result.errors.extend(report_result.errors)
            result.warnings.extend(report_result.warnings)
            result.passed.extend(report_result.passed)
        
        # Update is_valid flag
        result.is_valid = len(result.errors) == 0
        
        return result


if __name__ == "__main__":
    # Example usage
    print("Output Validator")
    print("=" * 50)
    print("This script validates calculations and logical consistency")
    print("Import this module and use OutputValidator class in your analysis")
