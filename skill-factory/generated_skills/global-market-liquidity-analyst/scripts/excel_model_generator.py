#!/usr/bin/env python3
"""
Excel Model Generator

Creates comprehensive Excel workbooks with liquidity calculations, formulas, 
charts, and dashboards. Ensures all calculations are transparent and auditable.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional
from datetime import datetime
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference as ChartReference
    from openpyxl.chart.axis import DateAxis
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")
import warnings
warnings.filterwarnings('ignore')

try:
    from output_validator import OutputValidator, ValidationResult
    VALIDATOR_AVAILABLE = True
except ImportError:
    VALIDATOR_AVAILABLE = False
    # Define dummy ValidationResult if not available
    class ValidationResult:
        def __init__(self):
            self.errors = []
            self.warnings = []
            self.passed = []
            self.is_valid = True


class ExcelModelGenerator:
    """Generate Excel workbooks with liquidity analysis models"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_liquidity_model(
        self,
        liquidity_data: pd.DataFrame,
        cycle_analysis: Optional[Dict] = None,
        central_bank_data: Optional[pd.DataFrame] = None,
        output_path: str = 'Global_Liquidity_Analysis_Model.xlsx'
    ) -> str:
        """
        Create comprehensive liquidity analysis Excel workbook
        
        Args:
            liquidity_data: Main liquidity time series data
            cycle_analysis: Optional cycle analysis results
            central_bank_data: Optional central bank balance sheet data
            output_path: Path to save Excel file
            
        Returns:
            Path to created Excel file
        """
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl is required for Excel generation. Install with: pip install openpyxl")
            return ""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create all required sheets (matching SKILL.md Step 10 specification)
            self._create_liquidity_cycle_sheet(wb, liquidity_data, cycle_analysis)
            
            if central_bank_data is not None:
                self._create_central_bank_sheet(wb, central_bank_data)
            
            # Create additional required sheets
            self._create_cross_border_flow_sheet(wb, liquidity_data)
            self._create_monetary_aggregates_sheet(wb, liquidity_data)
            self._create_asset_price_correlation_sheet(wb, liquidity_data)
            
            self._create_dashboard_sheet(wb, liquidity_data, cycle_analysis)
            
            # Create validation sheet
            if VALIDATOR_AVAILABLE:
                self._create_validation_sheet(wb, liquidity_data, cycle_analysis, central_bank_data)
            
            # Validate before saving
            validation_result = self.validate_excel_model(wb, liquidity_data, cycle_analysis, central_bank_data)
            
            # Save workbook
            wb.save(output_path)
            
            # Print validation summary
            if validation_result:
                if validation_result.is_valid:
                    print(f"✓ Excel model created and validated: {output_path}")
                else:
                    print(f"⚠ Excel model created with {len(validation_result.errors)} error(s): {output_path}")
                    for error in validation_result.errors[:3]:
                        print(f"  - {error['check']}: {error['message']}")
            else:
                print(f"✓ Excel model created: {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"Error creating Excel model: {e}")
            return ""
    
    def _create_liquidity_cycle_sheet(
        self,
        wb: Workbook,
        liquidity_data: pd.DataFrame,
        cycle_analysis: Optional[Dict]
    ):
        """Create liquidity cycle tracking worksheet"""
        ws = wb.create_sheet("Liquidity Cycle", 0)
        
        # Headers
        headers = ['Date', 'Liquidity Index', 'YoY Growth %', 'MoM Growth %', 'Cycle Phase', 'Cycle Completion %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 2
        for idx, data_row in liquidity_data.iterrows():
            ws.cell(row=row, column=1, value=data_row['date'] if 'date' in data_row else data_row.iloc[0])
            ws.cell(row=row, column=2, value=data_row.get('liquidity_index', data_row.iloc[1] if len(data_row) > 1 else 0))
            
            # Calculate growth rates
            if row > 2:
                prev_value = ws.cell(row=row-1, column=2).value
                current_value = ws.cell(row=row, column=2).value
                if prev_value and current_value and prev_value != 0:
                    # MoM growth formula: =(B3/B2-1)*100
                    ws.cell(row=row, column=4, value=f"=({get_column_letter(2)}{row}/{get_column_letter(2)}{row-1}-1)*100")
            
            # Add cycle phase if available
            if cycle_analysis and 'current_phase' in cycle_analysis:
                phase = cycle_analysis['current_phase'].get('phase', '')
                ws.cell(row=row, column=5, value=phase)
                completion = cycle_analysis['current_phase'].get('cycle_completion_percent', 0)
                ws.cell(row=row, column=6, value=completion)
            
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        
        # Add chart
        chart = LineChart()
        chart.title = "Liquidity Cycle"
        chart.style = 10
        chart.y_axis.title = 'Liquidity Index'
        chart.x_axis.title = 'Date'
        
        data = ChartReference(ws, min_col=2, min_row=1, max_row=row-1)
        cats = ChartReference(ws, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "H2")
    
    def _create_central_bank_sheet(
        self,
        wb: Workbook,
        central_bank_data: pd.DataFrame
    ):
        """Create central bank balance sheet analysis worksheet"""
        ws = wb.create_sheet("Central Banks", 1)
        
        # Headers
        headers = ['Date', 'Central Bank', 'Total Assets', 'MoM Change', 'YoY Change %', 'Policy Stance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 2
        for idx, data_row in central_bank_data.iterrows():
            ws.cell(row=row, column=1, value=data_row.get('date', data_row.iloc[0]))
            ws.cell(row=row, column=2, value=data_row.get('central_bank', 'Unknown'))
            ws.cell(row=row, column=3, value=data_row.get('total_assets', 0))
            
            row += 1
        
        # Format columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_cross_border_flow_sheet(
        self,
        wb: Workbook,
        liquidity_data: pd.DataFrame
    ):
        """Create cross-border capital flow tracking worksheet"""
        ws = wb.create_sheet("Cross-Border Flows", 3)
        
        # Headers
        headers = ['Date', 'Capital Flow', 'FX Liquidity Index', 'Reserve Currency Holdings', 'Swap Line Usage', 'Flow Direction']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data structure (placeholder - would need actual cross-border data)
        row = 2
        if len(liquidity_data) > 0:
            for idx, data_row in liquidity_data.iterrows():
                ws.cell(row=row, column=1, value=data_row.get('date', data_row.iloc[0] if len(data_row) > 0 else ''))
                # Placeholder columns - would be populated with actual cross-border flow data
                row += 1
        
        # Format columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Add note about data requirements
        note_row = row + 2
        ws.cell(row=note_row, column=1, value="Note: This sheet requires cross-border capital flow data from BIS, IMF, or central banks").font = Font(italic=True, size=9)
    
    def _create_monetary_aggregates_sheet(
        self,
        wb: Workbook,
        liquidity_data: pd.DataFrame
    ):
        """Create monetary aggregates tracking worksheet"""
        ws = wb.create_sheet("Monetary Aggregates", 4)
        
        # Headers
        headers = ['Date', 'M0 (Base Money)', 'M1 (Narrow Money)', 'M2 (Broad Money)', 'M3 (Extended)', 'M2 YoY Growth %', 'Velocity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data structure (placeholder - would need actual monetary aggregates data)
        row = 2
        if len(liquidity_data) > 0:
            for idx, data_row in liquidity_data.iterrows():
                ws.cell(row=row, column=1, value=data_row.get('date', data_row.iloc[0] if len(data_row) > 0 else ''))
                # Placeholder columns - would be populated with actual M0, M1, M2, M3 data
                row += 1
        
        # Format columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Add note about data requirements
        note_row = row + 2
        ws.cell(row=note_row, column=1, value="Note: This sheet requires M0, M1, M2, M3 data from central banks and statistical agencies").font = Font(italic=True, size=9)
    
    def _create_asset_price_correlation_sheet(
        self,
        wb: Workbook,
        liquidity_data: pd.DataFrame
    ):
        """Create asset price correlation analysis worksheet"""
        ws = wb.create_sheet("Asset Price Correlation", 5)
        
        # Headers
        headers = ['Date', 'Liquidity Index', 'Equity Index', 'Bond Yield', 'FX Rate', 'Correlation (Liquidity-Equity)', 'Correlation (Liquidity-Bonds)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data structure (placeholder - would need actual asset price data)
        row = 2
        if len(liquidity_data) > 0:
            for idx, data_row in liquidity_data.iterrows():
                ws.cell(row=row, column=1, value=data_row.get('date', data_row.iloc[0] if len(data_row) > 0 else ''))
                liquidity_val = data_row.get('liquidity_index', data_row.iloc[1] if len(data_row) > 1 else 0)
                ws.cell(row=row, column=2, value=liquidity_val)
                # Placeholder columns - would be populated with actual asset price data
                row += 1
        
        # Add correlation matrix section
        matrix_start_row = row + 3
        ws.cell(row=matrix_start_row, column=1, value="Correlation Matrix").font = Font(bold=True, size=12)
        matrix_start_row += 1
        
        correlation_labels = ['Liquidity', 'Equities', 'Bonds', 'FX']
        for i, label in enumerate(correlation_labels, 1):
            ws.cell(row=matrix_start_row, column=i+1, value=label).font = Font(bold=True)
            ws.cell(row=matrix_start_row+i, column=1, value=label).font = Font(bold=True)
            # Placeholder for correlation values
        
        # Format columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Add note about data requirements
        note_row = matrix_start_row + len(correlation_labels) + 2
        ws.cell(row=note_row, column=1, value="Note: This sheet requires asset price data (equities, bonds, FX) to calculate correlations").font = Font(italic=True, size=9)
    
    def _create_dashboard_sheet(
        self,
        wb: Workbook,
        liquidity_data: pd.DataFrame,
        cycle_analysis: Optional[Dict]
    ):
        """Create summary dashboard worksheet"""
        ws = wb.create_sheet("Dashboard", 2)
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws.cell(row=1, column=1, value="Global Market Liquidity Analysis Dashboard")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Key Metrics
        row = 3
        ws.cell(row=row, column=1, value="Key Metrics").font = Font(bold=True, size=12)
        row += 1
        
        metrics = [
            ["Current Liquidity Index", liquidity_data.iloc[-1].get('liquidity_index', 'N/A') if len(liquidity_data) > 0 else 'N/A'],
            ["Analysis Date", datetime.now().strftime('%Y-%m-%d')]
        ]
        
        if cycle_analysis and 'current_phase' in cycle_analysis:
            phase_info = cycle_analysis['current_phase']
            metrics.extend([
                ["Current Cycle Phase", phase_info.get('phase', 'Unknown')],
                ["Cycle Completion %", f"{phase_info.get('cycle_completion_percent', 0):.2f}%"],
                ["Months Elapsed", f"{phase_info.get('months_elapsed', 0):.2f}"]
            ])
        
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Format
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def create_comprehensive_model(
        self,
        data_dict: Dict[str, pd.DataFrame],
        analysis_results: Dict,
        output_path: str = 'comprehensive_liquidity_model.xlsx'
    ) -> str:
        """
        Create comprehensive workbook with multiple analysis components
        
        Args:
            data_dict: Dictionary of dataframes (liquidity, central_banks, etc.)
            analysis_results: Dictionary of analysis results
            output_path: Output file path
            
        Returns:
            Path to created Excel file
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Create sheets for each data type
            for sheet_name, df in data_dict.items():
                ws = wb.create_sheet(sheet_name.replace('_', ' ').title())
                self._write_dataframe_to_sheet(ws, df)
            
            # Save
            wb.save(output_path)
            print(f"✓ Comprehensive Excel model created: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"Error creating comprehensive model: {e}")
            return ""
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """Write DataFrame to Excel worksheet with formatting"""
        # Headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        # Data
        for row_idx, (_, data_row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def validate_excel_model(
        self,
        wb: Workbook,
        liquidity_data: Optional[pd.DataFrame] = None,
        cycle_analysis: Optional[Dict] = None,
        central_bank_data: Optional[pd.DataFrame] = None
    ) -> Optional[ValidationResult]:
        """
        Validate Excel model formulas and calculations
        
        Args:
            wb: Workbook object
            liquidity_data: Optional liquidity data for validation
            cycle_analysis: Optional cycle analysis results
            central_bank_data: Optional central bank data
            
        Returns:
            ValidationResult object or None if validator not available
        """
        if not VALIDATOR_AVAILABLE or not OPENPYXL_AVAILABLE:
            return None
        
        result = ValidationResult()
        validator = OutputValidator()
        
        try:
            # Check for division by zero errors in formulas
            division_by_zero_found = False
            
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Validation':
                    continue  # Skip validation sheet itself
                
                ws = wb[sheet_name]
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            formula = str(cell.value)
                            
                            # Check for division by zero
                            if '/0' in formula or 'DIV/0' in formula.upper():
                                division_by_zero_found = True
                                result.add_error(
                                    'excel_formula_validation',
                                    f'Potential division by zero in {sheet_name}!{cell.coordinate}: {formula}'
                                )
                
                # Validate percent change formulas in Liquidity Cycle sheet
                if sheet_name == 'Liquidity Cycle' and liquidity_data is not None:
                    # Check MoM growth formula (column D)
                    for row_idx in range(3, ws.max_row + 1):
                        mom_cell = ws.cell(row=row_idx, column=4)
                        if mom_cell.data_type == 'f':
                            formula = str(mom_cell.value)
                            # Verify formula structure: should be something like =B3/B2-1 or similar
                            if 'B' in formula and '/' in formula:
                                # Try to extract the calculation
                                prev_row = row_idx - 1
                                if prev_row >= 2:
                                    prev_value = ws.cell(row=prev_row, column=2).value
                                    curr_value = ws.cell(row=row_idx, column=2).value
                                    
                                    if prev_value and curr_value and prev_value != 0:
                                        expected_mom = ((curr_value - prev_value) / prev_value) * 100
                                        # Check if calculated value matches expected (within tolerance)
                                        if mom_cell.value is not None:
                                            try:
                                                # Evaluate the formula result if possible
                                                calculated_mom = mom_cell.value * 100 if isinstance(mom_cell.value, (int, float)) else None
                                                if calculated_mom is not None:
                                                    diff = abs(calculated_mom - expected_mom)
                                                    if diff > 1.0:  # 1% tolerance
                                                        result.add_warning(
                                                            'excel_mom_formula',
                                                            f'MoM growth formula in {sheet_name}!{mom_cell.coordinate} may be incorrect (diff: {diff:.2f}%)'
                                                        )
                                            except:
                                                pass
            
            if not division_by_zero_found:
                result.add_passed('excel_formula_validation', 'No division by zero errors found in formulas')
            
        except Exception as e:
            result.add_error('excel_model_validation', f'Error validating Excel model: {str(e)}')
        
        result.is_valid = len(result.errors) == 0
        return result
    
    def _create_validation_sheet(
        self,
        wb: Workbook,
        liquidity_data: Optional[pd.DataFrame],
        cycle_analysis: Optional[Dict],
        central_bank_data: Optional[pd.DataFrame]
    ):
        """Create validation worksheet with validation results"""
        if not VALIDATOR_AVAILABLE:
            return
        
        ws = wb.create_sheet("Validation", len(wb.sheetnames))
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws.cell(row=1, column=1, value="Validation Summary")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Run validation
        validator = OutputValidator()
        validation_results = []
        
        # Validate cycle analysis if available
        if cycle_analysis and liquidity_data is not None:
            cycle_result = validator.validate_cycle_phase_consistency(
                cycle_analysis,
                liquidity_data
            )
            validation_results.extend(cycle_result.errors)
            validation_results.extend(cycle_result.warnings)
            validation_results.extend(cycle_result.passed)
        
        # Headers
        row = 3
        headers = ['Check Type', 'Status', 'Check Name', 'Message']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Write validation results
        for check in validation_results:
            check_type = 'Error' if check in validation_results and 'error' in str(check).lower() else \
                        'Warning' if 'warning' in str(check).lower() else 'Passed'
            
            status_icon = '✗' if check_type == 'Error' else '⚠' if check_type == 'Warning' else '✓'
            check_name = check.get('check', 'Unknown')
            message = check.get('message', '')
            
            ws.cell(row=row, column=1, value=check_type)
            ws.cell(row=row, column=2, value=status_icon)
            ws.cell(row=row, column=3, value=check_name)
            ws.cell(row=row, column=4, value=message)
            
            # Color code based on status
            fill_color = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid") if check_type == 'Error' else \
                        PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid") if check_type == 'Warning' else \
                        PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Summary
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        summary_cell = ws.cell(row=row, column=1, value="Validation Summary: See checks above")
        summary_cell.font = Font(bold=True, size=12)
        
        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 60


if __name__ == "__main__":
    # Example usage
    print("Excel Model Generator")
    print("=" * 50)
    print("This script creates Excel workbooks with liquidity analysis")
    print("Import this module and use ExcelModelGenerator class in your analysis")
    
    # Note: Requires openpyxl library
    # Install with: pip install openpyxl
